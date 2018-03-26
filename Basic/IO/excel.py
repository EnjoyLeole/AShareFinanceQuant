import openpyxl as xl
from openpyxl.styles import Font


@property
def title():
    return Font(
        name = 'Arial',
        size = 24,
        bold = True,
        italic = True,
        vertAlign = None,
        underline = 'none',
        strike = False,
        color = 'FFD700')


class ExcelHelper(object):
    @property
    def blank(self):
        bk = []
        sep = " "
        for i in range(5):
            bk.append(sep)
        return bk

    def __init__(self, work_book_name = "test.xlsm"):
        super().__init__()
        self.file = work_book_name
        self.workbook = xl.load_workbook(self.file, keep_vba = True)

    def get_sheet(self, sheet = "report", if_exist = "append"):
        if self.workbook.sheetnames.__contains__(sheet):
            ws = self.workbook[sheet]
            if if_exist == "replace":
                self.sheet_clean(ws)
                ws._current_row = 0
            else:
                pass
        else:
            ws = self.workbook.create_sheet(sheet)
        return ws

    def df_to_excel(self, df, sheet = "report", if_exist = "append", index = False,
                    column_before = False):
        ws = self.get_sheet(sheet, if_exist)
        keys = df.keys().values.tolist()
        if column_before:
            keys.insert(0, "")
        if index:
            keys.insert(0, "")
        ws.append(keys)
        # print(df.shape[0])
        for i in range(df.shape[0]):
            li = df.ix[i].tolist()
            if index:
                li.insert(0, df.index[i])
            if column_before:
                li.insert(0, "")
            # print(list)
            ws.append(li)
        ws.append(self.blank)

    def series_to_excel(self, series, sheet = "report", if_exist = "append"):
        ws = self.get_sheet(sheet, if_exist)
        # ws = wb.active
        ws.append(series.keys().values.tolist())

        ws.append(series.values.tolist())
        ws.append(self.blank)

    @staticmethod
    def sheet_clean(sheet):
        for row in sheet.rows:
            for cell in row:
                cell.value = None

    def save(self):
        self.workbook.save(self.file)
